import io
import csv
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.models.invoice import StructuredDocument

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FONT = Font(bold=True, size=12, color="1E3A5F")
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

# ── Preferred column order for common line item fields ────────────────────────
# When GPT returns line items we enforce this display order.
# Any fields not in this list go at the end in whatever order they arrive.
PREFERRED_LINE_ITEM_ORDER = [
    "description", "quantity", "unit_price", "unit_cost",
    "rate", "hours", "amount", "total", "subtotal",
    "tax", "discount", "sku", "product_code", "variant",
]


def export_to_excel(doc: StructuredDocument) -> bytes:
    """
    Generates a formatted Excel file.
    Sheet 1: Summary with document info and totals.
    One sheet per section GPT extracted.
    """
    wb = Workbook()
    _build_summary_sheet(wb, doc)

    for name, data in doc.sections.items():
        title = name.replace("_", " ").title()[:31]
        if isinstance(data, list) and data:
            _build_table_sheet(wb, title, data)
        elif isinstance(data, dict) and data:
            _build_keyvalue_sheet(wb, title, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"Excel export: {len(doc.sections)} section sheets")
    return buf.getvalue()


def export_to_csv(doc: StructuredDocument) -> bytes:
    """
    Generates a CSV with all sections separated by headers.
    """
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["InvoiAI Export"])
    w.writerow(["Document Type", doc.document_type])
    w.writerow(["Industry", doc.industry])
    w.writerow(["Confidence", doc.extraction_confidence or ""])
    w.writerow([])

    if doc.raw_totals:
        w.writerow(["=== TOTALS ==="])
        for k, v in doc.raw_totals.items():
            w.writerow([_label(k), v])
        w.writerow([])

    for name, data in doc.sections.items():
        w.writerow([f"=== {name.upper().replace('_', ' ')} ==="])
        if isinstance(data, list) and data and isinstance(data[0], dict):
            keys = _ordered_keys(data)
            w.writerow([_label(k) for k in keys])
            for row in data:
                w.writerow([row.get(k, "") for k in keys])
        elif isinstance(data, dict):
            for k, v in data.items():
                w.writerow([_label(k), v if v is not None else ""])
        w.writerow([])

    return buf.getvalue().encode("utf-8")


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, doc: StructuredDocument) -> None:
    ws = wb.active
    ws.title = "Summary"
    row = 1

    # Title
    ws.cell(row=row, column=1, value="DOCUMENT SUMMARY").font = SECTION_FONT
    row += 1

    # Document info rows
    for label, value in [
        ("Document Type", doc.document_type),
        ("Industry", doc.industry),
        ("Confidence", doc.extraction_confidence or ""),
        ("Sections", len(doc.sections)),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1  # blank row

    # Totals block
    if doc.raw_totals:
        ws.cell(row=row, column=1, value="TOTALS").font = SECTION_FONT
        row += 1

        # Header row for totals
        for col, val in [(1, "Field"), (2, "Amount")]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left")
        row += 1

        for k, v in doc.raw_totals.items():
            ws.cell(row=row, column=1, value=_label(k))
            ws.cell(row=row, column=2, value=v)
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


def _build_table_sheet(wb: Workbook, title: str, data: list) -> None:
    """
    Table layout for list sections (line items, findings, transactions).
    Uses PREFERRED_LINE_ITEM_ORDER to put columns in the right order.
    """
    ws = wb.create_sheet(title=title)
    if not data or not isinstance(data[0], dict):
        return

    # Get keys in preferred order
    keys = _ordered_keys(data)

    # Header row
    for ci, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=ci, value=_label(k))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    # Data rows with alternating fill
    for ri, row in enumerate(data, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, k in enumerate(keys, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(k))
            if fill:
                cell.fill = fill

    # Auto-size columns
    for i, k in enumerate(keys, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(
            max(len(_label(k)) + 4, 12), 50
        )


def _build_keyvalue_sheet(wb: Workbook, title: str, data: dict) -> None:
    """
    Key-value layout for dict sections (header info, party details).
    Two columns: Field | Value.
    """
    ws = wb.create_sheet(title=title)

    # Header
    for col, val in [(1, "Field"), (2, "Value")]:
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    row = 2
    for k, v in data.items():
        if isinstance(v, dict):
            # Nested dict — sub-header then its fields
            ws.cell(row=row, column=1, value=_label(k)).font = Font(bold=True, italic=True)
            row += 1
            for sk, sv in v.items():
                ws.cell(row=row, column=1, value=f"  {_label(sk)}")
                ws.cell(row=row, column=2, value=sv)
                row += 1
        elif isinstance(v, list):
            ws.cell(row=row, column=1, value=_label(k))
            ws.cell(row=row, column=2, value=", ".join(str(i) for i in v))
            row += 1
        else:
            ws.cell(row=row, column=1, value=_label(k))
            ws.cell(row=row, column=2, value=v if v is not None else "")
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


# ── Helpers ───────────────────────────────────────────────────────────────────

def _ordered_keys(data: list[dict]) -> list[str]:
    """
    Returns column keys in preferred display order.

    Collects all unique keys from all rows first (some rows may have
    different fields). Then sorts them: preferred fields first in the
    order defined in PREFERRED_LINE_ITEM_ORDER, then any remaining
    fields alphabetically at the end.
    """
    all_keys: list[str] = []
    seen: set = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # Sort by preferred order
    preferred = [k for k in PREFERRED_LINE_ITEM_ORDER if k in seen]
    remaining = sorted([k for k in all_keys if k not in PREFERRED_LINE_ITEM_ORDER])
    return preferred + remaining


def _label(key: str) -> str:
    """Converts snake_case to Title Case for display."""
    return key.replace("_", " ").title()